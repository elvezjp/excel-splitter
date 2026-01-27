"""手動テスト用Excelファイルを作成する."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# デフォルト出力先
DEFAULT_OUTPUT = Path(__file__).parent / "fixtures" / "manual_test.xlsx"


def create_manual_test_file(filename: str | Path = DEFAULT_OUTPUT):
    """手動テスト用の包括的なテストファイルを作成する."""
    wb = Workbook()

    # ===================
    # 1. 目次シート
    # ===================
    ws_index = wb.active
    ws_index.title = "目次"

    ws_index["A1"] = "手動テスト用ファイル"
    ws_index["A1"].font = Font(bold=True, size=14)

    ws_index["A3"] = "リンク先"
    ws_index["B3"] = "説明"
    ws_index["A3"].font = Font(bold=True)
    ws_index["B3"].font = Font(bold=True)

    # スタイルシートへのリンク
    ws_index["A4"] = "スタイル確認シート"
    ws_index["A4"].hyperlink = "#スタイル確認!A1"
    ws_index["A4"].font = Font(color="0000FF", underline="single")
    ws_index["B4"] = "書式の維持を確認"

    # 大量データシートへのリンク（先頭）
    ws_index["A5"] = "大量データ（先頭）"
    ws_index["A5"].hyperlink = "#大量データ!A1"
    ws_index["A5"].font = Font(color="0000FF", underline="single")
    ws_index["B5"] = "PART1 に含まれる"

    # 大量データシートへのリンク（中盤 - 分割後はPART2）
    ws_index["A6"] = "大量データ 80行目"
    ws_index["A6"].hyperlink = "#大量データ!A80"
    ws_index["A6"].font = Font(color="0000FF", underline="single")
    ws_index["B6"] = "PART2 に含まれる（--max-rows 50 の場合）"

    # 大量データシートへのリンク（終盤 - 分割後はPART3）
    ws_index["A7"] = "大量データ 130行目"
    ws_index["A7"].hyperlink = "#大量データ!A130"
    ws_index["A7"].font = Font(color="0000FF", underline="single")
    ws_index["B7"] = "PART3 に含まれる（--max-rows 50 の場合）"

    ws_index.column_dimensions["A"].width = 25
    ws_index.column_dimensions["B"].width = 45

    # ===================
    # 2. スタイル確認シート
    # ===================
    ws_styles = wb.create_sheet("スタイル確認")

    # タイトル
    ws_styles["A1"] = "スタイル確認シート"
    ws_styles["A1"].font = Font(bold=True, size=14)

    # フォントスタイルセクション
    ws_styles["A3"] = "フォント:"
    ws_styles["A3"].font = Font(bold=True)

    ws_styles["A4"] = "太字テキスト"
    ws_styles["A4"].font = Font(bold=True)

    ws_styles["A5"] = "赤色テキスト"
    ws_styles["A5"].font = Font(color="FF0000")

    ws_styles["A6"] = "大きいテキスト（サイズ16）"
    ws_styles["A6"].font = Font(size=16)

    ws_styles["A7"] = "斜体テキスト"
    ws_styles["A7"].font = Font(italic=True)

    # 背景色セクション
    ws_styles["C3"] = "背景色:"
    ws_styles["C3"].font = Font(bold=True)

    ws_styles["C4"] = "黄色"
    ws_styles["C4"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws_styles["C5"] = "水色"
    ws_styles["C5"].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    ws_styles["C6"] = "薄緑"
    ws_styles["C6"].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # 罫線セクション
    ws_styles["E3"] = "罫線:"
    ws_styles["E3"].font = Font(bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

    ws_styles["E4"] = "細い罫線"
    ws_styles["E4"].border = thin_border

    ws_styles["E5"] = "太い罫線"
    ws_styles["E5"].border = thick_border

    # 格子罫線 (2x2)
    for row in range(6, 8):
        for col in range(5, 7):
            cell = ws_styles.cell(row=row, column=col)
            cell.value = f"格子 {row-5},{col-4}"
            cell.border = thin_border

    # セル結合セクション
    ws_styles["A10"] = "セル結合:"
    ws_styles["A10"].font = Font(bold=True)

    ws_styles.merge_cells("A11:B12")
    ws_styles["A11"] = "結合セル（2x2）"
    ws_styles["A11"].alignment = Alignment(horizontal="center", vertical="center")
    ws_styles["A11"].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    ws_styles.merge_cells("C11:E11")
    ws_styles["C11"] = "横方向結合（1x3）"
    ws_styles["C11"].alignment = Alignment(horizontal="center")

    # 列幅・行高セクション
    ws_styles["A15"] = "列幅・行高:"
    ws_styles["A15"].font = Font(bold=True)

    ws_styles["A16"] = "通常幅"
    ws_styles["B16"] = "広い列（幅30）"
    ws_styles["C16"] = "狭い列（幅15）"

    ws_styles.column_dimensions["A"].width = 20
    ws_styles.column_dimensions["B"].width = 30
    ws_styles.column_dimensions["C"].width = 15
    ws_styles.column_dimensions["E"].width = 15

    ws_styles["A18"] = "通常の行"
    ws_styles["A19"] = "高い行（高さ40）"
    ws_styles.row_dimensions[19].height = 40
    ws_styles["A19"].alignment = Alignment(vertical="center")

    # 目次へ戻るリンク
    ws_styles["A21"] = "目次へ戻る"
    ws_styles["A21"].hyperlink = "#目次!A1"
    ws_styles["A21"].font = Font(color="0000FF", underline="single")

    # ===================
    # 3. 大量データシート（複数テーブル）
    # ===================
    ws_large = wb.create_sheet("大量データ")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    current_row = 1

    # -----------------------
    # テーブル1: 売上データ（50行）- PART1に含まれる
    # -----------------------
    ws_large.cell(row=current_row, column=1, value="【テーブル1】売上データ")
    ws_large.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1

    # ヘッダー
    headers1 = ["売上ID", "商品名", "金額", "数量", "リンク"]
    for col, header in enumerate(headers1, 1):
        cell = ws_large.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border
    current_row += 1

    # データ行（48行）
    products = ["商品A", "商品B", "商品C", "商品D"]
    for i in range(1, 49):
        row = current_row
        ws_large.cell(row=row, column=1, value=f"S{i:03d}").border = thin_border
        ws_large.cell(row=row, column=2, value=products[i % 4]).border = thin_border
        ws_large.cell(row=row, column=3, value=i * 1000).border = thin_border
        ws_large.cell(row=row, column=4, value=(i % 10) + 1).border = thin_border

        if i % 2 == 0:
            for col in range(1, 5):
                ws_large.cell(row=row, column=col).fill = alt_row_fill

        current_row += 1

    # テーブル1内リンク（80行目へ → テーブル2に飛ぶ）
    ws_large.cell(row=10, column=5, value="80行目へ")
    ws_large.cell(row=10, column=5).hyperlink = "#大量データ!A80"
    ws_large.cell(row=10, column=5).font = Font(color="0000FF", underline="single")

    # 空白行
    current_row += 2

    # -----------------------
    # テーブル2: 在庫データ（50行）- PART2に含まれる
    # -----------------------
    table2_start = current_row
    ws_large.cell(row=current_row, column=1, value="【テーブル2】在庫データ")
    ws_large.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1

    # ヘッダー
    headers2 = ["在庫ID", "倉庫", "商品名", "在庫数", "リンク"]
    for col, header in enumerate(headers2, 1):
        cell = ws_large.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = thin_border
    current_row += 1

    # データ行（48行）
    warehouses = ["東京", "大阪", "名古屋", "福岡"]
    for i in range(1, 49):
        row = current_row
        ws_large.cell(row=row, column=1, value=f"Z{i:03d}").border = thin_border
        ws_large.cell(row=row, column=2, value=warehouses[i % 4]).border = thin_border
        ws_large.cell(row=row, column=3, value=products[i % 4]).border = thin_border
        ws_large.cell(row=row, column=4, value=i * 10).border = thin_border

        if i % 2 == 0:
            for col in range(1, 5):
                ws_large.cell(row=row, column=col).fill = alt_row_fill

        current_row += 1

    # テーブル2内リンク（10行目へ戻る → テーブル1に飛ぶ）
    ws_large.cell(row=80, column=5, value="10行目へ")
    ws_large.cell(row=80, column=5).hyperlink = "#大量データ!A10"
    ws_large.cell(row=80, column=5).font = Font(color="0000FF", underline="single")

    # テーブル2内リンク（130行目へ → テーブル3に飛ぶ）
    ws_large.cell(row=81, column=5, value="130行目へ")
    ws_large.cell(row=81, column=5).hyperlink = "#大量データ!A130"
    ws_large.cell(row=81, column=5).font = Font(color="0000FF", underline="single")

    # 空白行
    current_row += 2

    # -----------------------
    # テーブル3: 顧客データ（50行）- PART3に含まれる
    # -----------------------
    table3_start = current_row
    ws_large.cell(row=current_row, column=1, value="【テーブル3】顧客データ")
    ws_large.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1

    # ヘッダー
    headers3 = ["顧客ID", "氏名", "地域", "会員ランク", "リンク"]
    for col, header in enumerate(headers3, 1):
        cell = ws_large.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.border = thin_border
    current_row += 1

    # データ行（48行）
    regions = ["関東", "関西", "中部", "九州"]
    ranks = ["ゴールド", "シルバー", "ブロンズ", "一般"]
    for i in range(1, 49):
        row = current_row
        ws_large.cell(row=row, column=1, value=f"C{i:03d}").border = thin_border
        ws_large.cell(row=row, column=2, value=f"顧客{i:03d}").border = thin_border
        ws_large.cell(row=row, column=3, value=regions[i % 4]).border = thin_border
        ws_large.cell(row=row, column=4, value=ranks[i % 4]).border = thin_border

        if i % 2 == 0:
            for col in range(1, 5):
                ws_large.cell(row=row, column=col).fill = alt_row_fill

        current_row += 1

    # テーブル3内リンク（目次へ戻る）
    ws_large.cell(row=130, column=5, value="目次へ戻る")
    ws_large.cell(row=130, column=5).hyperlink = "#目次!A1"
    ws_large.cell(row=130, column=5).font = Font(color="0000FF", underline="single")

    # 列幅設定
    ws_large.column_dimensions["A"].width = 12
    ws_large.column_dimensions["B"].width = 15
    ws_large.column_dimensions["C"].width = 12
    ws_large.column_dimensions["D"].width = 12
    ws_large.column_dimensions["E"].width = 12

    # 保存
    wb.save(filename)
    print(f"手動テスト用ファイルを作成しました: {filename}")
    print()
    print("ファイル構成:")
    print("  - 目次: 他シートへのリンク")
    print("  - スタイル確認: フォント、背景色、罫線、結合セル、列幅/行高のサンプル")
    print("  - 大量データ: 3つのテーブル（売上/在庫/顧客）を含む")
    print("                --max-rows 50 で3つのPARTに分割される")


if __name__ == "__main__":
    create_manual_test_file()
