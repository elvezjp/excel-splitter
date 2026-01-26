from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def create_complex_test_file(filename: str):
    wb = Workbook()
    
    # 1. Main Sheet (Index)
    ws_main = wb.active
    ws_main.title = "Index_Sheet"
    
    # Add some headers
    ws_main["A1"] = "Test Scenario"
    ws_main["B1"] = "Link"
    ws_main["A1"].font = Font(bold=True)
    
    # Link to Sheet 2
    ws_main["A2"] = "Go to Sales Data"
    ws_main["B2"] = "Link"
    ws_main["B2"].hyperlink = "#'Sales 2024'!A1"
    ws_main["B2"].font = Font(color="0000FF", underline="single")
    
    # Link to Sheet 3 (Large)
    ws_main["A3"] = "Go to Large Data"
    ws_main["B3"] = "Link"
    ws_main["B3"].hyperlink = "#'Large_Data'!A1"
    ws_main["B3"].font = Font(color="0000FF", underline="single")

    # Link to Special Char Sheet
    ws_main["A4"] = "Special Chars"
    ws_main["B4"] = "Link"
    ws_main["B4"].hyperlink = "#'A:B<C>'!A1" # Invalid chars usually forbidden in Excel GUI but via code...
    # Actually Excel forbids : \ ? * [ ] / in sheet names.
    # We should use characters that ARE allowed but need sanitization for FS?
    # Valid in Excel but maybe tricky: Space, Single Quote, etc.
    # Let's try "Sheet With Spaces"
    
    # 2. Sales Data (Normal Sheet)
    ws_sales = wb.create_sheet("Sales 2024")
    ws_sales["A1"] = "Revenue"
    ws_sales["B1"] = 10000
    ws_sales.sheet_properties.tabColor = "1072BA" # Set color
    
    # Link back to Index
    ws_sales["C1"] = "Back to Index"
    ws_sales["C1"].hyperlink = "#Index_Sheet!A1"
    
    # 3. Large Data (For row splitting)
    # We will use --max-rows 50 for testing, so let's make 120 rows.
    ws_large = wb.create_sheet("Large_Data")
    ws_large["A1"] = "ID"
    ws_large["B1"] = "Value"
    ws_large["C1"] = "Description"
    
    for i in range(1, 121):
        row = i + 1
        ws_large.cell(row=row, column=1, value=i)
        ws_large.cell(row=row, column=2, value=f"Val_{i}")
        ws_large.cell(row=row, column=3, value="Some description text")
        
        # Style every 10th row
        if i % 10 == 0:
            ws_large.cell(row=row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(filename)
    print(f"Created test file: {filename}")

if __name__ == "__main__":
    create_complex_test_file("complex_test.xlsx")
