import os
import shutil
import pytest
from openpyxl import load_workbook
from click.testing import CliRunner
from excel_splitter.cli import main
from tests.create_test_data import create_complex_test_file

# Constants
TEST_INPUT = "complex_test.xlsx"
TEST_OUTPUT_DIR = "test_output"

@pytest.fixture(scope="module")
def setup_test_files():
    """Setup and teardown test files"""
    # Setup
    create_complex_test_file(TEST_INPUT)
    if os.path.exists(TEST_OUTPUT_DIR):
        shutil.rmtree(TEST_OUTPUT_DIR)
    
    yield
    
    # Teardown
    if os.path.exists(TEST_INPUT):
        os.remove(TEST_INPUT)
    if os.path.exists(TEST_OUTPUT_DIR):
        shutil.rmtree(TEST_OUTPUT_DIR)


def test_cli_basic_execution(setup_test_files):
    """Test that CLI executes without error"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR])
    
    assert result.exit_code == 0, f"CLI should succeed. Output: {result.output}"


def test_cli_with_max_rows(setup_test_files):
    """Test CLI with --max-rows option"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR, "--max-rows", "50"])
    
    assert result.exit_code == 0, f"CLI with --max-rows should succeed. Output: {result.output}"


def test_cli_dry_run(setup_test_files):
    """Test CLI --dry-run flag"""
    runner = CliRunner()
    dry_run_dir = TEST_OUTPUT_DIR + "_dryrun"
    result = runner.invoke(main, [TEST_INPUT, "-o", dry_run_dir, "--dry-run"])

    assert result.exit_code == 0
    assert "[Dry Run]" in result.output
    # Directory should NOT be created by dry-run
    assert not os.path.exists(dry_run_dir)


def test_cli_verbose(setup_test_files):
    """Test CLI --verbose flag"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR, "--verbose"])
    
    assert result.exit_code == 0
    assert "Input:" in result.output
    assert "Output:" in result.output


def test_cli_invalid_input_file():
    """Test CLI with non-existent input file"""
    runner = CliRunner()
    result = runner.invoke(main, ["nonexistent.xlsx", "-o", TEST_OUTPUT_DIR])
    
    assert result.exit_code != 0


def test_cli_non_xlsx_file():
    """Test CLI with non-.xlsx file"""
    runner = CliRunner()
    
    # Create a temporary non-.xlsx file
    temp_file = "test_file.txt"
    with open(temp_file, "w") as f:
        f.write("test")
    
    try:
        result = runner.invoke(main, [temp_file, "-o", TEST_OUTPUT_DIR])
        assert result.exit_code != 0
        assert "Only .xlsx files" in result.output or "Error" in result.output
    finally:
        if os.path.exists(temp_file):
            os.remove(temp_file)


def test_workbook_split_creates_files(setup_test_files):
    """Test that workbook split creates individual sheet files"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR])
    
    assert result.exit_code == 0
    
    # Check that split files were created
    index_file = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Index_Sheet.xlsx")
    sales_file = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Sales_2024.xlsx")
    
    assert os.path.exists(index_file), "Index sheet file should be created"
    assert os.path.exists(sales_file), "Sales sheet file should be created"


def test_row_split_creates_parts(setup_test_files):
    """Test that row split creates part files"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR, "--max-rows", "50"])
    
    assert result.exit_code == 0
    
    # Large_Data sheet should be split into parts (not exist as single file)
    large_data_file = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Large_Data.xlsx")
    
    # Should be split into parts instead
    part1 = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Large_Data_PART1.xlsx")
    part2 = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Large_Data_PART2.xlsx")
    part3 = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Large_Data_PART3.xlsx")
    
    assert os.path.exists(part1), "Part 1 should be created"
    assert os.path.exists(part2), "Part 2 should be created"
    assert os.path.exists(part3), "Part 3 should be created"


def test_row_split_content(setup_test_files):
    """Test that row split creates parts with correct row counts"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR, "--max-rows", "50"])

    assert result.exit_code == 0

    part1 = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Large_Data_PART1.xlsx")
    part2 = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Large_Data_PART2.xlsx")
    part3 = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Large_Data_PART3.xlsx")

    # Part 1: rows 1-50
    wb_p1 = load_workbook(part1)
    ws_p1 = wb_p1.active
    assert ws_p1.max_row == 50, "Part 1 should have 50 rows"
    wb_p1.close()

    # Part 2: rows 51-100
    wb_p2 = load_workbook(part2)
    ws_p2 = wb_p2.active
    assert ws_p2["A1"].value is not None, "Part 2 should have data"
    assert ws_p2.max_row == 50, "Part 2 should have 50 rows"
    wb_p2.close()

    # Part 3: rows 101-121 (21 rows)
    wb_p3 = load_workbook(part3)
    ws_p3 = wb_p3.active
    assert ws_p3["A1"].value is not None, "Part 3 should have data"
    assert ws_p3.max_row == 21, "Part 3 should have 21 rows"
    wb_p3.close()


def test_hyperlink_rewriting(setup_test_files):
    """Test that internal hyperlinks are rewritten"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR])
    
    assert result.exit_code == 0
    
    index_file = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Index_Sheet.xlsx")
    sales_file = os.path.join(TEST_OUTPUT_DIR, "complex_test__SHEET__Sales_2024.xlsx")
    
    # Check hyperlinks in Index sheet
    wb_index = load_workbook(index_file)
    ws_index = wb_index.active
    
    # Cell B2 should have a hyperlink
    link_b2 = ws_index.cell(row=2, column=2).hyperlink
    assert link_b2 is not None, "Cell B2 should have a hyperlink"
    
    target = link_b2.target
    assert "./" in target, "Link should be a relative external path"
    assert "complex_test__SHEET__Sales_2024.xlsx" in target, "Link should reference correct file"
    
    wb_index.close()


def test_sanitized_filenames(setup_test_files):
    """Test that special characters in sheet names are sanitized"""
    runner = CliRunner()
    result = runner.invoke(main, [TEST_INPUT, "-o", TEST_OUTPUT_DIR])
    
    assert result.exit_code == 0
    
    # Verify that the output directory contains properly named files
    output_files = os.listdir(TEST_OUTPUT_DIR)
    
    # All files should be .xlsx
    for file in output_files:
        assert file.endswith(".xlsx"), f"File {file} should be .xlsx"
        # Should not contain illegal filename characters
        illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
        for char in illegal_chars:
            assert char not in file, f"File {file} contains illegal character: {char}"


def test_output_directory_creation(setup_test_files):
    """Test that output directory is created if it doesn't exist"""
    runner = CliRunner()
    
    non_existent_dir = os.path.join(TEST_OUTPUT_DIR, "subdir", "nested")
    result = runner.invoke(main, [TEST_INPUT, "-o", non_existent_dir])
    
    # Should still succeed (output dir should be created)
    assert result.exit_code == 0
    assert os.path.exists(non_existent_dir), "Output directory should be created"
    
    # Cleanup
    if os.path.exists(non_existent_dir):
        shutil.rmtree(os.path.dirname(non_existent_dir))
