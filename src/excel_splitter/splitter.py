import copy
import os
from openpyxl import load_workbook, Workbook

def split_workbook_by_sheet(input_path: str, output_dir: str, verbose: bool = False) -> list[tuple[str, str]]:
    """
    Split a workbook into multiple files, one per sheet.
    Uses 'Delete Other Sheets' method to preserve styles.
    
    Returns:
        List of (sheet_name, output_file_path) tuples.
    """
    if verbose:
        print(f"Loading workbook: {input_path} (This might take a while for large files...)")
    
    # We load once to get sheet names, but for 'delete other sheets' properly without 
    # side effects between iterations, we might need to reload or copy carefully.
    # openpyxl objects are not easily deepcopyable in all cases.
    # Strategy: storage of original wb in memory is tricky if we delete sheets.
    # Simplest Robust Strategy: Load, delete others, save. Repeat.
    # This is slow O(N^2) load time if we reload every time.
    # Optimized Strategy: Load once into memory as a template? openpyxl copy is hard.
    # Let's stick to the Spec's "Accuracy First" approach -> Reload or iterative delete.
    # Actually, we can load ONCE, then for each sheet, we might need a copy.
    # But deepcopying a large Workbook is also slow/problematic.
    
    # Let's try: Load master. For each sheet to export:
    # 1. Check if we can duplicate the file on disk? No, that's just IO.
    # 2. To be safe and simple for MVP:
    #    Iterate sheet names. For each sheet:
    #      Load WB.
    #      Remove all other sheets.
    #      Save.
    # This is standard "Delete Other Sheets".
    
    # Optimization: Read sheet names effectively first.
    # 'read_only=True' to just get names?
    temp_wb = load_workbook(input_path, read_only=True)
    sheet_names = temp_wb.sheetnames
    temp_wb.close()
    
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    generated_files = []

    from .utils import get_sheet_filename

    for target_sheet in sheet_names:
        if verbose:
            print(f"Processing sheet: {target_sheet}")
            
        # Refreshed load for each iteration to ensure clean state
        # data_only=False to keep formulas (though refs might break, we want the formula text)
        wb = load_workbook(input_path)
        
        # Remove others
        for sheet in wb.sheetnames:
            if sheet != target_sheet:
                del wb[sheet]
        
        # Determine output path
        filename = get_sheet_filename(base_name, target_sheet)
        out_path = os.path.join(output_dir, filename)
        
        wb.save(out_path)
        wb.close()
        generated_files.append((target_sheet, out_path))
        
    return generated_files
