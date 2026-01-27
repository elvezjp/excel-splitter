import os
from copy import copy
from openpyxl import Workbook, load_workbook

from .utils import get_part_filename
from .hyperlinks import (
    extract_internal_link,
    generate_external_link,
    generate_part_external_link,
    get_range_part_decision_from_boundaries,
    get_part_number_from_boundaries,
)

def is_blank_row(ws, row_num: int) -> bool:
    """
    Check if a row is blank (all cells have value None).
    """
    for cell in ws[row_num]:
        if cell.value is not None:
            return False
    return True


def is_all_blank_rows(ws, start_row: int, end_row: int) -> bool:
    """
    Check if all rows in the range [start_row, end_row] are blank.
    """
    for row_num in range(start_row, end_row + 1):
        if not is_blank_row(ws, row_num):
            return False
    return True


def compute_all_boundaries(ws, max_rows: int, total_rows: int) -> list[tuple[int, int]]:
    """
    Pre-compute all part boundaries for a sheet.
    Returns list of (start_row, end_row) tuples for each part.
    Skips entirely blank ranges.
    """
    boundaries = []
    current_row = 1

    while current_row <= total_rows:
        max_end_row = min(current_row + max_rows - 1, total_rows)
        end_row = find_split_point(ws, current_row, max_end_row, total_rows)

        # Skip if all rows in this range are blank
        if not is_all_blank_rows(ws, current_row, end_row):
            boundaries.append((current_row, end_row))

        current_row = end_row + 1

    return boundaries


def find_split_point(ws, start_row: int, max_end_row: int, total_rows: int) -> int:
    """
    Find the best split point within the range [start_row, max_end_row].

    Strategy:
    1. Search for the last blank row in range (start_row, max_end_row]
    2. If found, split before the blank row (return blank_row - 1)
    3. If not found, use max_end_row as fallback

    Note: We exclude start_row from blank row search to avoid infinite loops
    when start_row itself is a blank row.
    """
    search_end = min(max_end_row, total_rows)

    # Find the last blank row in range (excluding start_row)
    last_blank_row = None
    for row_num in range(start_row + 1, search_end + 1):
        if is_blank_row(ws, row_num):
            last_blank_row = row_num

    if last_blank_row is not None:
        # Split before the blank row
        return last_blank_row - 1
    else:
        # Fallback: use search_end
        return search_end


def copy_row_style(source_cell, target_cell):
    """
    Copy basic style and hyperlink from source to target.
    Deepcopying style objects is safer than assigning references if across workbooks.
    """
    if source_cell.has_style:
        # Assigning style objects directly can work if openpyxl version supports it, 
        # but often copying attributes is safer for different workbooks.
        # For simplicity/speed in 'large sheet' context, we might only copy values if too slow.
        # Let's try to copy essentials: font, fill, border, alignment.
        try:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
        except Exception:
            pass
    
    # Copy hyperlink if present
    if source_cell.hyperlink:
        try:
            target_cell.hyperlink = copy(source_cell.hyperlink)
        except Exception:
            # If hyperlink copy fails, log but don't crash
            pass


def split_sheet_by_rows(
    origin_wb_path: str,
    sheet_name: str,
    output_dir: str,
    max_rows: int,
    base_name: str,
    split_map: dict[str, list[tuple[int, int]]],
    verbose: bool = False
) -> tuple[list[str], list[tuple[int, int]]]:
    """
    Split a specific sheet into multiple files based on max_rows.
    Returns tuple of (list of generated file paths, list of part boundaries).
    Each boundary is (start_row, end_row) tuple.
    """
    
    # Load source. Read-only might be faster for reading, 
    # but we need cell styles. So standard load.
    wb_source = load_workbook(origin_wb_path, read_only=False, data_only=False)
    ws_source = wb_source[sheet_name]
    
    # Count rows
    total_rows = ws_source.max_row

    # If total_rows <= max_rows, no split needed
    if total_rows <= max_rows:
        wb_source.close()
        return [], [] # No split happened here. Caller handles "Single file" case.

    # Pre-compute all part boundaries for this sheet
    part_boundaries = compute_all_boundaries(ws_source, max_rows, total_rows)

    generated_files = []

    for part_num, (current_row, end_row) in enumerate(part_boundaries, 1):
        if verbose:
            print(f"  Creating Part {part_num} for {sheet_name} (Row {current_row}-{end_row})")

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        row_write_idx = 1
        for row in ws_source.iter_rows(min_row=current_row, max_row=end_row, values_only=False):
            for col_idx, cell in enumerate(row, 1):
                new_cell = new_ws.cell(row=row_write_idx, column=col_idx, value=cell.value)
                copy_row_style(cell, new_cell)

                # Rewrite internal links that point to a different part of the same sheet,
                # or to other sheets.
                if cell.hyperlink:
                    link = extract_internal_link(cell.hyperlink)
                    if link:
                        if link.sheet_name == sheet_name:
                            # Same sheet: use pre-computed boundaries
                            target_part, ref_for_link, spans_parts = get_range_part_decision_from_boundaries(link, part_boundaries)
                            if target_part != part_num or spans_parts:
                                new_target = generate_part_external_link(base_name, sheet_name, ref_for_link, target_part)
                                new_cell.hyperlink = new_target
                                if verbose and spans_parts:
                                    print(f"  [Link Rewrite] Range spans parts at {cell.coordinate}: {link.ref} -> {ref_for_link}")
                        else:
                            # Link to another sheet: use split_map boundaries
                            target_boundaries = split_map.get(link.sheet_name, [])
                            if target_boundaries:
                                target_part, ref_for_link, spans_parts = get_range_part_decision_from_boundaries(link, target_boundaries)
                                new_target = generate_part_external_link(base_name, link.sheet_name, ref_for_link, target_part)
                                if verbose and spans_parts:
                                    print(f"  [Link Rewrite] Range spans parts at {cell.coordinate}: {link.ref} -> {ref_for_link}")
                            else:
                                new_target = generate_external_link(base_name, link.sheet_name, link.ref)
                            new_cell.hyperlink = new_target
            row_write_idx += 1
            
        # Save
        filename = get_part_filename(base_name, sheet_name, part_num)
        out_path = os.path.join(output_dir, filename)
        new_wb.save(out_path)
        new_wb.close()

        generated_files.append(out_path)

    wb_source.close()
    return generated_files, part_boundaries
