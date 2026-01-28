import os
import sys
import click
from openpyxl import load_workbook

from .splitter import split_workbook_by_sheet
from .hyperlinks import rewrite_hyperlinks_zip, has_internal_links
from .row_splitter import split_sheet_by_rows, compute_all_boundaries
from .utils import get_sheet_filename, get_part_filename

@click.command()
@click.argument('input_file', type=click.Path(exists=True, dir_okay=False))
@click.option('-o', '--output-dir', default='./dist', help='Directory to save split files.', type=click.Path())
@click.option('--max-rows', default=0, help='Maximum rows per sheet. Splits into parts if exceeded.', type=int)
@click.option('--dry-run', is_flag=True, help='Simulate split without writing files.')
@click.option('--verbose', is_flag=True, help='Enable verbose output.')
def main(input_file, output_dir, max_rows, dry_run, verbose):
    """
    Excel Splitter: Splits .xlsx files into individual sheets or row chunks.
    Preserves hyperlinks by rewriting them to external relative links.
    Preserves shapes/images when possible (no internal links, no row splitting).
    """
    try:
        if verbose:
            click.echo(f"Input: {input_file}")
            click.echo(f"Output: {output_dir}")
            click.echo(f"Max Rows: {max_rows}")
            click.echo("---")

        # Validation: .xlsx extension check
        if not input_file.lower().endswith('.xlsx'):
            click.echo("Error: Only .xlsx files are supported.", err=True)
            sys.exit(1)

        # Validation: Check if input file is readable
        if not os.access(input_file, os.R_OK):
            click.echo(f"Error: Cannot read input file: {input_file}", err=True)
            sys.exit(1)

        # Validation: max_rows should be non-negative
        if max_rows < 0:
            click.echo("Error: --max-rows must be >= 0", err=True)
            sys.exit(1)

        base_name = os.path.splitext(os.path.basename(input_file))[0]

        # Validate output directory path (avoid file path)
        if os.path.exists(output_dir) and not os.path.isdir(output_dir):
            click.echo(f"Error: Output path is not a directory: {output_dir}", err=True)
            sys.exit(1)

        # DRY RUN LOGIC
        if dry_run:
            click.echo("[Dry Run] Simulating split...")
            try:
                wb = load_workbook(input_file, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    total_rows = ws.max_row
                    if max_rows > 0 and total_rows > max_rows:
                        parts = (total_rows + max_rows - 1) // max_rows
                        click.echo(f"- Would split sheet: {sheet} ({total_rows} rows) -> {parts} parts")
                        for part_num in range(1, parts + 1):
                            filename = get_part_filename(base_name, sheet, part_num)
                            click.echo(f"  - {filename}")
                    else:
                        filename = get_sheet_filename(base_name, sheet)
                        click.echo(f"- Would create file for sheet: {sheet} -> {filename}")
                wb.close()
            except Exception as e:
                click.echo(f"Error: Failed to read workbook: {e}", err=True)
                sys.exit(1)
            return

        # REAL RUN
        # Create output directory with proper error handling
        try:
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
        except PermissionError:
            click.echo(f"Error: No write permission for directory: {output_dir}", err=True)
            sys.exit(1)
        except Exception as e:
            click.echo(f"Error: Failed to create output directory: {e}", err=True)
            sys.exit(1)

        # Validate write permission
        if not os.access(output_dir, os.W_OK):
            click.echo(f"Error: No write permission for directory: {output_dir}", err=True)
            sys.exit(1)

        # Step 1: ZIP-based Workbook Split (preserves shapes/images)
        click.echo(">> Phase 1: Splitting sheets (ZIP operation)...")
        try:
            generated_files = split_workbook_by_sheet(input_file, output_dir, verbose)
        except Exception as e:
            click.echo(f"Error: Phase 1 failed (Workbook Split): {e}", err=True)
            sys.exit(1)

        # Step 2: Analyze original file with openpyxl
        # - Count rows for each sheet
        # - Check for internal links
        # - Identify sheets that need row splitting
        click.echo(">> Phase 2: Analyzing workbook...")
        split_map = {}  # {sheet_name: [(start_row, end_row), ...]}
        sheets_with_internal_links = set()  # シートに内部リンクがあるかどうか

        try:
            wb_meta = load_workbook(input_file, read_only=False, data_only=False)
            for sheet_name in wb_meta.sheetnames:
                ws = wb_meta[sheet_name]
                total_rows = ws.max_row

                # Check for row splitting need
                if max_rows > 0 and total_rows > max_rows:
                    boundaries = compute_all_boundaries(ws, max_rows, total_rows)
                    if len(boundaries) > 1:
                        split_map[sheet_name] = boundaries
                        if verbose:
                            click.echo(f"  Sheet '{sheet_name}': {len(boundaries)} parts needed")

                # Check for internal links
                if has_internal_links(ws):
                    sheets_with_internal_links.add(sheet_name)
                    if verbose:
                        click.echo(f"  Sheet '{sheet_name}': has internal links")

            wb_meta.close()
        except Exception as e:
            click.echo(f"Error: Failed to analyze workbook: {e}", err=True)
            sys.exit(1)

        # Step 3: Conditional post-processing for each sheet
        click.echo(">> Phase 3: Post-processing sheets...")

        final_file_list = []

        for sheet_name, file_path in generated_files:
            try:
                needs_row_split = sheet_name in split_map
                needs_link_rewrite = sheet_name in sheets_with_internal_links

                if needs_row_split:
                    # Row splitting (openpyxl) - shapes will be lost
                    if verbose:
                        click.echo(f"  [Row Split] Sheet '{sheet_name}' exceeds {max_rows} rows. Splitting...")

                    try:
                        part_files, _ = split_sheet_by_rows(
                            input_file, sheet_name, output_dir, max_rows, base_name, split_map, verbose
                        )
                    except Exception as e:
                        click.echo(f"Error: Row split failed for sheet '{sheet_name}': {e}", err=True)
                        sys.exit(1)

                    if part_files:
                        # Delete the whole sheet file to avoid duplication
                        try:
                            os.remove(file_path)
                        except Exception as e:
                            click.echo(f"Warning: Failed to delete original file after split: {e}")
                        final_file_list.extend(part_files)
                        click.echo(f"  -> Created {len(part_files)} parts for '{sheet_name}'")
                    else:
                        final_file_list.append(file_path)

                elif needs_link_rewrite:
                    # Hyperlink rewriting (ZIP/XML) - shapes/images are preserved
                    if verbose:
                        click.echo(f"  [Link Rewrite] Sheet '{sheet_name}' has internal links. Rewriting (ZIP/XML)...")

                    rewrite_hyperlinks_zip(file_path, base_name, sheet_name, split_map, verbose)
                    final_file_list.append(file_path)

                else:
                    # No post-processing needed
                    if verbose:
                        click.echo(f"  [No Change] Sheet '{sheet_name}'")
                    final_file_list.append(file_path)

            except Exception as e:
                click.echo(f"Error: Post-processing failed for sheet '{sheet_name}': {e}", err=True)
                sys.exit(1)

        click.echo(f"\nDone. Generated {len(final_file_list)} files in {output_dir}")

    except Exception as e:
        click.echo(f"Unexpected error: {e}", err=True)
        sys.exit(1)

if __name__ == '__main__':
    main()
